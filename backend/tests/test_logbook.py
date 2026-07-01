"""Vehicle logbook: vehicle CRUD, trip CRUD, chronological journey-number
renumbering, xlsx export, and the AI trip-suggest endpoint."""
import io
import types

from openpyxl import load_workbook

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _vehicle(client, auth_headers, **over):
    body = {"ecv": "BA123AB", "consumption": 6.5, "fuel_price": 1.6}
    body.update(over)
    r = client.post("/api/vehicles", json=body, headers=auth_headers)
    assert r.status_code == 201, r.text
    return r.json()


def _trip(client, auth_headers, vid, start_dt, **over):
    body = {
        "start_dt": start_dt,
        "purpose": "Nákup tovaru",
        "route": "Bratislava > Trnava > Bratislava",
        "km": 120,
    }
    body.update(over)
    r = client.post(f"/api/vehicles/{vid}/trips", json=body, headers=auth_headers)
    assert r.status_code == 201, r.text
    return r.json()


def _trips_for_month(client, auth_headers, vid, year, month):
    r = client.get(
        f"/api/vehicles/{vid}/trips", params={"year": year, "month": month}, headers=auth_headers
    )
    assert r.status_code == 200, r.text
    return r.json()


def _build_import_xlsx(rows):
    """A minimal xlsx matching the layout parse_xlsx() looks for: a header
    row containing "začiatku jazdy", then data rows with start_dt in column
    B and route in column E."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.cell(13, 2, "Dátum a čas začiatku jazdy")
    for i, row in enumerate(rows, start=14):
        ws.cell(i, 2, row["start"])
        ws.cell(i, 5, row.get("route", ""))
        ws.cell(i, 8, row.get("km", 0))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- Vehicles ----

def test_create_and_list_vehicle(client, auth_headers):
    v = _vehicle(client, auth_headers)
    r = client.get("/api/vehicles", headers=auth_headers)
    assert r.status_code == 200
    out = r.json()[0]
    assert out["id"] == v["id"]
    assert out["km_total"] is None  # no trips yet
    assert out["km_ytd"] is None


def test_vehicle_km_stats_split_by_year(client, auth_headers):
    v = _vehicle(client, auth_headers, odometer_base=1000)
    _trip(client, auth_headers, v["id"], "2025-12-20T09:00:00", km=50)
    _trip(client, auth_headers, v["id"], "2026-01-15T09:00:00", km=100)

    out = client.get("/api/vehicles", headers=auth_headers).json()[0]
    assert out["km_total"] == 1000 + 150
    assert out["km_ytd"] == 100  # only the 2026 trip counts towards this year


def test_delete_vehicle_blocked_while_it_has_trips(client, auth_headers):
    v = _vehicle(client, auth_headers)
    t = _trip(client, auth_headers, v["id"], "2026-07-05T09:00:00")

    r = client.delete(f"/api/vehicles/{v['id']}", headers=auth_headers)
    assert r.status_code == 409

    client.delete(f"/api/car-trips/{t['id']}", headers=auth_headers)
    r = client.delete(f"/api/vehicles/{v['id']}", headers=auth_headers)
    assert r.status_code == 204


def test_vehicles_require_auth(client):
    r = client.get("/api/vehicles")
    assert r.status_code == 401


# ---- Trip CRUD ----

def test_create_trip_assigns_first_journey_number(client, auth_headers):
    v = _vehicle(client, auth_headers)
    t = _trip(client, auth_headers, v["id"], "2026-07-05T09:00:00")
    assert t["journey_number"] == 202607001


def test_update_and_delete_trip(client, auth_headers):
    v = _vehicle(client, auth_headers)
    t = _trip(client, auth_headers, v["id"], "2026-07-05T09:00:00", purpose="Pôvodný účel")

    r = client.patch(f"/api/car-trips/{t['id']}", json={"purpose": "Nový účel"}, headers=auth_headers)
    assert r.status_code == 200
    assert r.json()["purpose"] == "Nový účel"

    r = client.delete(f"/api/car-trips/{t['id']}", headers=auth_headers)
    assert r.status_code == 204
    assert _trips_for_month(client, auth_headers, v["id"], 2026, 7) == []


def test_create_trip_for_missing_vehicle_404s(client, auth_headers):
    r = client.post(
        "/api/vehicles/999999/trips",
        json={"start_dt": "2026-07-05T09:00:00"},
        headers=auth_headers,
    )
    assert r.status_code == 404


# ---- Journey-number chronological renumbering ----
# Regression coverage for the bug where trips added out of date order kept
# their insertion-order journey number instead of a chronological one.

def test_creating_an_earlier_trip_renumbers_the_later_one(client, auth_headers):
    v = _vehicle(client, auth_headers)
    late = _trip(client, auth_headers, v["id"], "2026-07-20T09:00:00")
    assert late["journey_number"] == 202607001  # only trip so far

    early = _trip(client, auth_headers, v["id"], "2026-07-05T09:00:00")

    trips = _trips_for_month(client, auth_headers, v["id"], 2026, 7)
    by_id = {t["id"]: t for t in trips}
    assert by_id[early["id"]]["journey_number"] == 202607001
    assert by_id[late["id"]]["journey_number"] == 202607002


def test_deleting_a_trip_renumbers_the_remaining_ones(client, auth_headers):
    v = _vehicle(client, auth_headers)
    t1 = _trip(client, auth_headers, v["id"], "2026-07-05T09:00:00")
    t2 = _trip(client, auth_headers, v["id"], "2026-07-10T09:00:00")
    t3 = _trip(client, auth_headers, v["id"], "2026-07-15T09:00:00")

    client.delete(f"/api/car-trips/{t2['id']}", headers=auth_headers)

    trips = {t["id"]: t for t in _trips_for_month(client, auth_headers, v["id"], 2026, 7)}
    assert trips[t1["id"]]["journey_number"] == 202607001
    assert trips[t3["id"]]["journey_number"] == 202607002


def test_moving_a_trip_to_another_month_renumbers_both_months(client, auth_headers):
    v = _vehicle(client, auth_headers)
    t1 = _trip(client, auth_headers, v["id"], "2026-07-05T09:00:00")
    t2 = _trip(client, auth_headers, v["id"], "2026-07-10T09:00:00")
    t3 = _trip(client, auth_headers, v["id"], "2026-08-03T09:00:00")

    # Move t1 into August, dated before the existing August trip.
    r = client.patch(
        f"/api/car-trips/{t1['id']}", json={"start_dt": "2026-08-01T08:00:00"}, headers=auth_headers
    )
    assert r.status_code == 200

    july = {t["id"]: t for t in _trips_for_month(client, auth_headers, v["id"], 2026, 7)}
    assert list(july) == [t2["id"]]
    assert july[t2["id"]]["journey_number"] == 202607001  # gap closed

    august = {t["id"]: t for t in _trips_for_month(client, auth_headers, v["id"], 2026, 8)}
    assert august[t1["id"]]["journey_number"] == 202608001  # earlier date wins #1
    assert august[t3["id"]]["journey_number"] == 202608002


# ---- Import ----

def test_import_out_of_order_rows_renumbers_chronologically(client, auth_headers):
    v = _vehicle(client, auth_headers)
    data = _build_import_xlsx([
        {"start": "20.07.2026 09:00", "route": "A > B", "km": 50},
        {"start": "05.07.2026 09:00", "route": "C > D", "km": 30},
    ])
    r = client.post(
        f"/api/vehicles/{v['id']}/trips/import",
        files={"file": ("logbook.xlsx", data, _XLSX_MIME)},
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    assert r.json()["imported"] == 2

    trips = {t["route"]: t for t in _trips_for_month(client, auth_headers, v["id"], 2026, 7)}
    assert trips["C > D"]["journey_number"] == 202607001  # earlier date -> #1
    assert trips["A > B"]["journey_number"] == 202607002


# ---- Export ----

def test_export_month_bases_odometer_on_prior_km(client, auth_headers):
    v = _vehicle(client, auth_headers, odometer_base=500)
    _trip(client, auth_headers, v["id"], "2026-06-15T09:00:00", km=50)
    _trip(client, auth_headers, v["id"], "2026-07-10T09:00:00", km=80)

    r = client.get(
        f"/api/vehicles/{v['id']}/trips/export",
        params={"year": 2026, "month": 7},
        headers=auth_headers,
    )
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws.cell(14, 6).value == 550  # odometer start = base(500) + prior June km(50)
    assert ws.cell(14, 7).value == 630  # odometer end = start + this trip's 80 km


def test_export_without_year_month_covers_whole_logbook(client, auth_headers):
    v = _vehicle(client, auth_headers, odometer_base=500)
    _trip(client, auth_headers, v["id"], "2026-06-15T09:00:00", km=50)
    _trip(client, auth_headers, v["id"], "2026-07-10T09:00:00", km=80)

    r = client.get(f"/api/vehicles/{v['id']}/trips/export", headers=auth_headers)
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    # Two data rows, running odometer starting from odometer_base (no month filter).
    assert ws.cell(14, 6).value == 500 and ws.cell(14, 7).value == 550
    assert ws.cell(15, 6).value == 550 and ws.cell(15, 7).value == 630
    assert "attachment" in r.headers["content-disposition"]
    assert "_2026_" not in r.headers["content-disposition"]  # no month suffix


def test_export_404s_when_vehicle_has_no_trips(client, auth_headers):
    v = _vehicle(client, auth_headers)
    r = client.get(f"/api/vehicles/{v['id']}/trips/export", headers=auth_headers)
    assert r.status_code == 404


# ---- AI trip suggest ----

def _patch_anthropic(monkeypatch, text=None, error=None):
    import anthropic

    class _Messages:
        def create(self, **kwargs):
            if error is not None:
                raise error
            return types.SimpleNamespace(content=[types.SimpleNamespace(text=text)])

    class _Client:
        def __init__(self, api_key=None):
            self.messages = _Messages()

    monkeypatch.setattr(anthropic, "Anthropic", _Client)


def test_ai_trip_suggest_without_api_key_returns_503(client, auth_headers, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    r = client.post("/api/ai-trip-suggest", json={"description": "cesta na alzu"}, headers=auth_headers)
    assert r.status_code == 503


def test_ai_trip_suggest_parses_fenced_json_response(client, auth_headers, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    _patch_anthropic(
        monkeypatch,
        text='```json\n{"purpose": "Nákup", "route": "Bratislava > Trnava > Bratislava", '
        '"km": 80, "start_time": "08:00", "end_time": "12:00", "trip_type": "Firemná"}\n```',
    )
    r = client.post(
        "/api/ai-trip-suggest",
        json={"description": "cesta do Trnavy po tovar", "home_city": "Bratislava", "date": "2026-07-10"},
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["km"] == 80
    assert body["trip_type"] == "Firemná"


def test_ai_trip_suggest_coerces_bad_fields_instead_of_failing(client, auth_headers, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    _patch_anthropic(
        monkeypatch,
        text='{"purpose": "Nákup", "route": "Bratislava > Trnava", "km": "about 40", '
        '"start_time": "8:00", "end_time": "12:00", "trip_type": "Neviem"}',
    )
    r = client.post("/api/ai-trip-suggest", json={"description": "cesta na alzu"}, headers=auth_headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["km"] is None            # "about 40" isn't numeric -> coerced to None
    assert body["start_time"] is None    # "8:00" doesn't match HH:MM -> coerced to None
    assert body["end_time"] == "12:00"   # well-formed field passes through
    assert body["trip_type"] == "Firemná"  # unrecognized value -> falls back to default


def test_ai_trip_suggest_upstream_error_returns_502_with_message(client, auth_headers, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    _patch_anthropic(monkeypatch, error=RuntimeError("credit balance too low"))
    r = client.post("/api/ai-trip-suggest", json={"description": "cesta na alzu"}, headers=auth_headers)
    assert r.status_code == 502
    assert "credit balance too low" in r.json()["detail"]


def test_ai_trip_suggest_unparseable_response_returns_500(client, auth_headers, monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    _patch_anthropic(monkeypatch, text="Sorry, I can't help with that request.")
    r = client.post("/api/ai-trip-suggest", json={"description": "cesta na alzu"}, headers=auth_headers)
    assert r.status_code == 500
