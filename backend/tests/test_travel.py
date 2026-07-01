"""Travel report: per-diem bands, CRUD, and xlsx export."""
import io
from datetime import date, time
from decimal import Decimal

import openpyxl

from app import travel as tv

R = tv.DEFAULT_RATES
D = date(2026, 7, 1)


def test_per_diem_bands():
    # New signature: computed_per_diem(date_from, date_to, first_depart, last_arrive, rates)
    assert tv.computed_per_diem(D, None, time(9, 0), time(13, 0), R) == Decimal("0.00")    # 4h
    assert tv.computed_per_diem(D, None, time(7, 30), time(15, 30), R) == Decimal("8.80")  # 8h
    assert tv.computed_per_diem(D, None, time(7, 30), time(20, 30), R) == Decimal("13.10") # 13h
    assert tv.computed_per_diem(D, None, time(7, 0), time(6, 0), R) == Decimal("19.50")    # 23h overnight
    assert tv.computed_per_diem(D, None, None, None, R) == Decimal("0.00")
    # Multi-day: 1 Jul 08:00 -> 2 Jul 18:00 = 34h => 1 full day (19.50) + 10h (8.80).
    assert tv.computed_per_diem(D, date(2026, 7, 2), time(8, 0), time(18, 0), R) == Decimal("28.30")


def _period(client, auth_headers, year=2026, month=7):
    return client.post("/api/periods", json={"year": year, "month": month}, headers=auth_headers).json()["id"]


def _trip(client, auth_headers, pid, **over):
    body = {
        "traveller_name": "Nikoleta", "traveller_address": "Nitra",
        "trip_date": "2026-07-01", "purpose": "Konzultácia",
        "legs": [
            {"from_place": "Nitra", "to_place": "Trnava", "transport": "Auto služobné",
             "depart_time": "07:30", "arrive_time": "08:15"},
            {"from_place": "Trnava", "to_place": "Nitra", "transport": "Auto služobné",
             "depart_time": "14:45", "arrive_time": "15:30"},
        ],
    }
    body.update(over)
    return client.post(f"/api/periods/{pid}/travels", json=body, headers=auth_headers)


def test_create_list_and_per_diem(client, auth_headers):
    pid = _period(client, auth_headers)
    res = _trip(client, auth_headers, pid)
    assert res.status_code == 201, res.text
    body = res.json()
    assert body["per_diem"] == 8.8 and body["per_diem_computed"] == 8.8
    assert body["duration_hours"] == 8.0

    lst = client.get(f"/api/periods/{pid}/travels", headers=auth_headers).json()
    assert len(lst) == 1
    names = client.get(f"/api/periods/{pid}/travel-names", headers=auth_headers).json()
    assert names == ["Nikoleta"]


def test_override_and_clear(client, auth_headers):
    pid = _period(client, auth_headers, month=8)
    trip = _trip(client, auth_headers, pid).json()
    leg_id = trip["legs"][0]["id"]
    # Set per_diem on first leg — effective per_diem becomes sum of leg per_diems
    upd = client.patch(f"/api/travel-legs/{leg_id}", json={"per_diem": 20.0}, headers=auth_headers).json()
    assert upd["per_diem"] == 20.0 and upd["per_diem_computed"] == 8.8
    # Clear it — falls back to duration-based
    cleared = client.patch(f"/api/travel-legs/{leg_id}", json={"per_diem": None}, headers=auth_headers).json()
    assert cleared["per_diem"] == 8.8


def test_delete_trip(client, auth_headers):
    pid = _period(client, auth_headers, month=9)
    tid = _trip(client, auth_headers, pid).json()["id"]
    assert client.delete(f"/api/travels/{tid}", headers=auth_headers).status_code == 204
    assert client.get(f"/api/periods/{pid}/travels", headers=auth_headers).json() == []


def test_per_diem_rates_get_set(client, auth_headers):
    got = client.get("/api/travel/per-diem-rates", headers=auth_headers).json()
    assert got == {"band1": 8.8, "band2": 13.1, "band3": 19.5}
    client.patch("/api/travel/per-diem-rates", json={"band1": 9.0, "band2": 14.0, "band3": 21.0}, headers=auth_headers)
    # New rate flows into computed per-diem.
    pid = _period(client, auth_headers, month=10)
    body = _trip(client, auth_headers, pid).json()
    assert body["per_diem"] == 9.0


def test_export_xlsx(client, auth_headers):
    pid = _period(client, auth_headers, month=11)
    _trip(client, auth_headers, pid)
    # Second trip: 14h → band2 → 13.10
    _trip(client, auth_headers, pid, trip_date="2026-11-02", legs=[
        {"from_place": "Nitra", "to_place": "Trnava", "transport": "Auto služobné",
         "depart_time": "07:00", "arrive_time": "08:00"},
        {"from_place": "Trnava", "to_place": "Nitra", "transport": "Auto služobné",
         "depart_time": "20:00", "arrive_time": "21:00"},
    ])

    res = client.get(f"/api/periods/{pid}/travels/export", params={"name": "Nikoleta"}, headers=auth_headers)
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in res.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert wb.sheetnames == ["November", "VPC"]
    s1 = wb["November"]
    assert s1["B2"].value == "CESTOVNÝ PRÍKAZ"
    assert s1["C4"].value == "Nikoleta"

    vpc = wb["VPC"]
    spolu = [c.value for row in vpc.iter_rows() for c in row if c.value == "SPOLU"]
    assert spolu, "SPOLU row present"
    # Stravné (column F) are raw floats; SPOLU uses a formula so we sum the raw values.
    # Each trip's last Príchod row carries its effective per_diem: 8.80 + 13.10 = 21.90
    f_vals = [c.value for row in vpc.iter_rows() for c in row
              if c.column == 6 and isinstance(c.value, (int, float))]
    assert abs(sum(f_vals) - 21.90) < 0.001, f"expected stravné total 21.90, got {f_vals}"


def test_multiday_trip(client, auth_headers):
    pid = _period(client, auth_headers, month=3)
    res = _trip(client, auth_headers, pid, trip_date="2026-03-01", end_date="2026-03-02", legs=[
        {"from_place": "Nitra", "to_place": "Trnava", "transport": "Auto služobné",
         "depart_time": "08:00", "arrive_time": "09:00"},
        {"from_place": "Trnava", "to_place": "Nitra", "transport": "Auto služobné",
         "depart_time": "17:00", "arrive_time": "18:00"},
    ])
    body = res.json()
    assert body["end_date"] == "2026-03-02"
    assert body["per_diem"] == 28.3  # 34h -> 19.50 + 8.80


def test_duplicate_trip(client, auth_headers):
    pid = _period(client, auth_headers, month=4)
    tid = _trip(client, auth_headers, pid).json()["id"]
    dup = client.post(f"/api/travels/{tid}/duplicate", headers=auth_headers)
    assert dup.status_code == 201, dup.text
    assert dup.json()["id"] != tid
    assert len(client.get(f"/api/periods/{pid}/travels", headers=auth_headers).json()) == 2


def test_update_trip_date_moves_it_to_matching_period(client, auth_headers):
    march_pid = _period(client, auth_headers, month=3)
    april_pid = _period(client, auth_headers, month=4)
    tid = _trip(client, auth_headers, march_pid, trip_date="2026-03-15").json()["id"]

    res = client.patch(f"/api/travels/{tid}", json={"trip_date": "2026-04-10"}, headers=auth_headers)
    assert res.status_code == 200, res.text
    assert res.json()["period_id"] == april_pid

    assert client.get(f"/api/periods/{march_pid}/travels", headers=auth_headers).json() == []
    april_list = client.get(f"/api/periods/{april_pid}/travels", headers=auth_headers).json()
    assert [t["id"] for t in april_list] == [tid]


def test_update_trip_date_without_matching_period_404s_and_keeps_original(client, auth_headers):
    march_pid = _period(client, auth_headers, month=3)
    tid = _trip(client, auth_headers, march_pid, trip_date="2026-03-15").json()["id"]

    res = client.patch(f"/api/travels/{tid}", json={"trip_date": "2026-05-01"}, headers=auth_headers)
    assert res.status_code == 404
    assert "Months" in res.json()["detail"]

    unchanged = client.get(f"/api/periods/{march_pid}/travels", headers=auth_headers).json()
    assert len(unchanged) == 1 and unchanged[0]["trip_date"] == "2026-03-15"


def test_bulk_create_trips(client, auth_headers):
    pid = _period(client, auth_headers, month=5)
    body = {
        "traveller_name": "Bulk Person", "traveller_address": "Nitra",
        "trip_date": "2026-05-05", "purpose": "Konzultácia",
        "dates": ["2026-05-05", "2026-05-12", "2026-05-19"],
        "legs": [
            {"from_place": "Nitra", "to_place": "Trnava", "transport": "Vlak",
             "depart_time": "07:30", "arrive_time": "08:15"},
            {"from_place": "Trnava", "to_place": "Nitra", "transport": "Vlak",
             "depart_time": "14:45", "arrive_time": "15:30"},
        ],
    }
    res = client.post(f"/api/periods/{pid}/travels/bulk", json=body, headers=auth_headers)
    assert res.status_code == 201, res.text
    created = res.json()
    assert len(created) == 3
    assert sorted(t["trip_date"] for t in created) == ["2026-05-05", "2026-05-12", "2026-05-19"]
    assert all(t["per_diem"] == 8.8 for t in created)


def test_period_with_trips_cannot_be_deleted(client, auth_headers):
    pid = _period(client, auth_headers, month=6)
    _trip(client, auth_headers, pid)
    res = client.delete(f"/api/periods/{pid}", headers=auth_headers)
    assert res.status_code == 409  # must clear trips first (no FK 500)


def test_export_missing_person_404(client, auth_headers):
    pid = _period(client, auth_headers, month=12)
    res = client.get(f"/api/periods/{pid}/travels/export", params={"name": "Nobody"}, headers=auth_headers)
    assert res.status_code == 404
