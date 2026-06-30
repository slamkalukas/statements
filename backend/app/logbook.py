"""Vehicle logbook (Kniha jázd): xlsx export matching the company format."""
import io
from datetime import datetime

from .models import CarTrip, Vehicle
from .travel import COMPANY

_SK_MONTHS = {
    1: "Január", 2: "Február", 3: "Marec", 4: "Apríl", 5: "Máj", 6: "Jún",
    7: "Júl", 8: "August", 9: "September", 10: "Október", 11: "November", 12: "December",
}


def fuel_unit(vehicle: Vehicle) -> str:
    return "kWh" if "elektr" in (vehicle.fuel_type or "").lower() else "L"


def trip_cost(trip: CarTrip, vehicle: Vehicle) -> float | None:
    if trip.km is None or vehicle.consumption is None:
        return None
    price = float(trip.fuel_price_override or vehicle.fuel_price or 0)
    if not price:
        return None
    return round(trip.km * float(vehicle.consumption) / 100 * price, 2)


def build_xlsx(vehicle: Vehicle, trips: list[CarTrip], base_odometer: int = 0) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    unit = fuel_unit(vehicle)
    bold = Font(bold=True)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title = f"{vehicle.ecv} {vehicle.manufacturer} {vehicle.car_model}".strip()
    ws.title = title[:31]

    meta = [
        ("Kniha jázd", None),
        ("Firma", COMPANY),
        ("Dátum exportu", datetime.today().strftime("%d.%m.%Y")),
        ("Vlastníctvo vozidla", vehicle.ownership or "Firemné"),
        ("EČV", vehicle.ecv),
        ("VIN", vehicle.vin or ""),
        ("Výrobca", vehicle.manufacturer or ""),
        ("Model", vehicle.car_model or ""),
        ("Palivo", vehicle.fuel_type or ""),
        (f"Spotreba [{unit}/100km]",
         float(vehicle.consumption) if vehicle.consumption is not None else None),
        ("Dátum zaradenia do majetku",
         vehicle.date_added.strftime("%d.%m.%Y") if vehicle.date_added else ""),
        ("Spôsob výpočtu trás", "Spotreba podľa technického preukazu"),
    ]
    for r, (label, value) in enumerate(meta, start=1):
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = bold
        if value is not None:
            ws.cell(r, 2).value = value

    header_row = 13
    headers = [
        "Číslo záznamu o jazde",
        "Dátum a čas začiatku jazdy",
        "Dátum a čas skončenia jazdy",
        "Účel jazdy",
        "Jazda / cesta",
        "Počiatočný stav odometra",
        "Konečný stav odometra",
        "Počet najazdených km",
        f"Cena PHM [EUR/{unit}]",
        "Typ jazdy",
        "Vodič",
        "Udalosti",
        "Náklady na cestu [EUR]",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, c, value=h)
        cell.font = bold
        cell.border = box
        cell.alignment = center

    trips_sorted = sorted(trips, key=lambda t: t.start_dt)
    running = base_odometer
    for i, trip in enumerate(trips_sorted, start=1):
        r = header_row + i
        odo_start = running
        odo_end = running + (trip.km or 0)
        running = odo_end
        cost = trip_cost(trip, vehicle)
        fp = float(trip.fuel_price_override or vehicle.fuel_price or 0) or None
        row_data = [
            trip.journey_number,
            trip.start_dt.strftime("%d.%m.%Y %H:%M") if trip.start_dt else "",
            trip.end_dt.strftime("%d.%m.%Y %H:%M") if trip.end_dt else "",
            trip.purpose,
            trip.route,
            odo_start,
            odo_end,
            trip.km,
            fp,
            trip.trip_type,
            trip.driver_name,
            trip.events,
            cost,
        ]
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(r, c, value=val)
            cell.border = box

    widths = [15, 20, 20, 30, 50, 14, 14, 10, 12, 12, 20, 20, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- Import ----

def _parse_dt(v) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _int(v) -> int | None:
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _float(v) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    # Find header row dynamically — look for the start-time column header
    header_row = 13  # our default
    for row in ws.iter_rows(min_row=1, max_row=25):
        for cell in row:
            if isinstance(cell.value, str) and "začiatku jazdy" in cell.value:
                header_row = cell.row
                break
        else:
            continue
        break

    # Columns (0-based): 0=journey_num, 1=start_dt, 2=end_dt, 3=purpose, 4=route,
    # 5=odo_start, 6=odo_end, 7=km(skip), 8=fuel_price, 9=trip_type, 10=driver,
    # 11=events, 12=cost(skip)
    trips = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            break
        start_dt = _parse_dt(row[1] if len(row) > 1 else None)
        if start_dt is None:
            continue
        odo_start = _int(row[5]) if len(row) > 5 else None
        odo_end = _int(row[6]) if len(row) > 6 else None
        km = _int(row[7]) if len(row) > 7 else None
        if km is None and odo_start is not None and odo_end is not None:
            km = odo_end - odo_start
        trips.append({
            "start_dt": start_dt,
            "end_dt": _parse_dt(row[2] if len(row) > 2 else None),
            "purpose": str(row[3] or "") if len(row) > 3 else "",
            "route": str(row[4] or "") if len(row) > 4 else "",
            "km": km,
            "fuel_price_override": _float(row[8]) if len(row) > 8 else None,
            "trip_type": str(row[9] or "Firemná") if len(row) > 9 else "Firemná",
            "driver_name": str(row[10] or "") if len(row) > 10 else "",
            "events": (str(row[11]) if row[11] else None) if len(row) > 11 else None,
        })
    return trips
