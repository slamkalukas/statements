"""Travel report (cestovné): per-diem calculation and xlsx export.

Per-diem (stravné) follows Slovak meal-allowance bands by trip duration; for
multi-leg/international trips each leg can carry its own per_diem value and the
trip total is their sum. The export reproduces the two-sheet template:
"Cestovný príkaz" + "Vyúčtovanie pracovnej cesty" (VPC), per person per month.
"""
import io
import json
from datetime import date, time
from decimal import Decimal

from .models import Setting, Travel

RATES_KEY = "per_diem_rates"
# Slovak 2025 defaults: 5–12h, >12–18h, >18h. Below 5h -> 0.
DEFAULT_RATES = {"band1": 8.80, "band2": 13.10, "band3": 19.50}

COMPANY = "dotCUBE s.r.o"

_SK_MONTHS = {
    1: "Január", 2: "Február", 3: "Marec", 4: "Apríl", 5: "Máj", 6: "Jún",
    7: "Júl", 8: "August", 9: "September", 10: "Október", 11: "November", 12: "December",
}

_COMPANY_CAR_MARKERS = ("firemn", "služobn", "sluzob")


def is_company_car_transport(transport: str | None) -> bool:
    """True for transport labels that mean "company car" (Auto firemné/služobné),
    matched loosely enough to survive diacritics/spelling variants."""
    t = (transport or "").lower()
    return any(marker in t for marker in _COMPANY_CAR_MARKERS)


def get_rates(db) -> dict:
    row = db.query(Setting).filter(Setting.key == RATES_KEY).first()
    if row and row.value:
        try:
            data = json.loads(row.value)
            return {k: float(data.get(k, DEFAULT_RATES[k])) for k in DEFAULT_RATES}
        except (ValueError, TypeError):
            pass
    return dict(DEFAULT_RATES)


def set_rates(db, rates: dict) -> dict:
    clean = {k: round(float(rates.get(k, DEFAULT_RATES[k])), 2) for k in DEFAULT_RATES}
    row = db.query(Setting).filter(Setting.key == RATES_KEY).first()
    if row:
        row.value = json.dumps(clean)
    else:
        db.add(Setting(key=RATES_KEY, value=json.dumps(clean)))
    return clean


def duration_hours(date_from: date, date_to: date | None,
                   first_depart: time | None, last_arrive: time | None) -> float | None:
    """Trip length in hours: from first leg's departure to last leg's arrival home."""
    if first_depart is None or last_arrive is None:
        return None
    end = date_to or date_from
    span_days = (end - date_from).days
    d = first_depart.hour * 60 + first_depart.minute
    a = last_arrive.hour * 60 + last_arrive.minute
    total = span_days * 24 * 60 + (a - d)
    if total <= 0:
        total += 24 * 60
    return total / 60.0


def _band_amount(hours: float, rates: dict) -> float:
    if hours < 5:
        return 0.0
    if hours <= 12:
        return rates["band1"]
    if hours <= 18:
        return rates["band2"]
    return rates["band3"]


def computed_per_diem(date_from: date, date_to: date | None,
                      first_depart: time | None, last_arrive: time | None,
                      rates: dict) -> Decimal:
    """Per-diem from trip duration using the configured bands. For multi-day trips
    each full 24h counts as a whole-day allowance (band3) plus the remainder by band."""
    h = duration_hours(date_from, date_to, first_depart, last_arrive)
    if h is None:
        return Decimal("0.00")
    full_days = int(h // 24)
    remainder = h - full_days * 24
    amount = full_days * rates["band3"] + _band_amount(remainder, rates)
    return Decimal(str(amount)).quantize(Decimal("0.01"))


def _leg_times(t: Travel) -> tuple[time | None, time | None]:
    """Return (first leg depart_time, last leg arrive_time) for duration calculation."""
    if not t.legs:
        return None, None
    return t.legs[0].depart_time, t.legs[-1].arrive_time


def effective_per_diem(t: Travel, rates: dict) -> Decimal:
    """Sum of leg per_diems when any leg has one set; otherwise compute from duration."""
    if t.legs and any(leg.per_diem is not None for leg in t.legs):
        total = sum(
            Decimal(str(leg.per_diem)) for leg in t.legs if leg.per_diem is not None
        )
        return total.quantize(Decimal("0.01"))
    first_depart, last_arrive = _leg_times(t)
    return computed_per_diem(t.trip_date, t.end_date, first_depart, last_arrive, rates)


def _fmt_date(d: date) -> str:
    return f"{d.day}.{d.month}.{d.year}"


def _fmt_time(t: time | None) -> str:
    return f"{t.hour}:{t.minute:02d}" if t is not None else ""


def _home_place(t: Travel) -> str:
    return t.legs[0].from_place if t.legs else ""


def _meeting_places(t: Travel) -> str:
    seen: list[str] = []
    for leg in t.legs:
        if leg.to_place and leg.to_place not in seen:
            seen.append(leg.to_place)
    return ", ".join(seen)


def build_xlsx(name: str, address: str, year: int, month: int,
               travels: list[Travel], rates: dict) -> bytes:
    """Render the two-sheet travel report for one person and month."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    travels = sorted(travels, key=lambda t: (t.trip_date, t.legs[0].depart_time if t.legs else time(0, 0)))

    wb = Workbook()

    # ---- Sheet 1: Cestovný príkaz ----
    s1 = wb.active
    s1.title = _SK_MONTHS.get(month, str(month))
    s1["B2"] = "CESTOVNÝ PRÍKAZ"; s1["B2"].font = title_font
    s1["B3"] = "Firma:"; s1["C3"] = "Meno a priezvisko:"; s1["E3"] = "Bydlisko:"
    for c in ("B3", "C3", "E3"):
        s1[c].font = bold
    s1["B4"] = COMPANY; s1["C4"] = name; s1["E4"] = address

    s1["B6"] = "Začiatok cesty:"; s1["C6"] = "Miesto rokovania:"
    s1["E6"] = "Účel cesty:"; s1["G6"] = "Koniec cesty:"
    for c in ("B6", "C6", "E6", "G6"):
        s1[c].font = bold
    r = 7
    for t in travels:
        end = t.end_date or t.trip_date
        first_leg = t.legs[0] if t.legs else None
        last_leg = t.legs[-1] if t.legs else None
        s1[f"B{r}"] = f"{_fmt_date(t.trip_date)} {first_leg.from_place if first_leg else ''}, {_fmt_time(first_leg.depart_time if first_leg else None)}".strip(", ")
        s1[f"C{r}"] = _meeting_places(t)
        s1[f"E{r}"] = t.purpose
        s1[f"G{r}"] = f"{_fmt_date(end)} {last_leg.to_place if last_leg else ''}, {_fmt_time(last_leg.arrive_time if last_leg else None)}".strip(", ")
        for c in ("B", "C", "E", "G"):
            s1[f"{c}{r}"].border = box
        r += 1
    for col, w in {"B": 26, "C": 20, "D": 6, "E": 40, "F": 6, "G": 26}.items():
        s1.column_dimensions[col].width = w

    # ---- Sheet 2: Vyúčtovanie pracovnej cesty (VPC) ----
    # Columns: B=Dátum C=ODCHOD–PRÍCHOD D=o hod. E=Dopravný prostriedok
    #          F=Stravné  G=Výdavky  H=Spolu
    s2 = wb.create_sheet("VPC")
    s2["B2"] = "VYÚČTOVANIE PRACOVNEJ CESTY"; s2["B2"].font = title_font
    s2["B3"] = "Firma:"; s2["C3"] = "Meno a priezvisko:"
    s2["B3"].font = bold; s2["C3"].font = bold
    s2["B4"] = COMPANY; s2["C4"] = name

    headers = ["Dátum", "ODCHOD – PRÍCHOD", "o hod.", "Použitý dopravný prostriedok",
               "Stravné", "Výdavky", "Spolu"]
    for i, h in enumerate(headers):
        cell = s2.cell(row=6, column=2 + i, value=h)
        cell.font = bold; cell.alignment = center; cell.border = box

    data_start = 7
    r = data_start

    for t in travels:
        end = t.end_date or t.trip_date
        has_leg_per_diem = any(leg.per_diem is not None for leg in t.legs)
        trip_pd = float(effective_per_diem(t, rates)) if not has_leg_per_diem else None

        for i, leg in enumerate(t.legs):
            is_last_leg = (i == len(t.legs) - 1)
            # First leg → trip_date; last leg → end_date; middle legs → leg_date or trip_date
            if i == 0:
                effective_date = t.trip_date
            elif is_last_leg:
                effective_date = end
            else:
                effective_date = leg.leg_date or t.trip_date
            leg_date = _fmt_date(effective_date)
            arrive_date = leg_date

            # Row 1: Odchod from_place at depart_time
            s2.cell(r, 2).value = leg_date
            s2.cell(r, 3).value = f"Odchod {leg.from_place}".strip()
            s2.cell(r, 4).value = _fmt_time(leg.depart_time)
            s2.cell(r, 5).value = leg.transport
            for col in range(2, 9):
                s2.cell(r, col).border = box
            r += 1

            # Row 2: Príchod to_place at arrive_time — expense/per_diem go here
            s2.cell(r, 2).value = arrive_date
            s2.cell(r, 3).value = f"Príchod {leg.to_place}".strip()
            s2.cell(r, 4).value = _fmt_time(leg.arrive_time)
            # Per-diem: per-leg if set; otherwise show trip total on last Príchod row only
            pd_value = None
            if leg.per_diem is not None:
                pd_value = float(leg.per_diem)
            elif is_last_leg and trip_pd is not None:
                pd_value = trip_pd
            if pd_value is not None:
                s2.cell(r, 6).value = pd_value
                s2.cell(r, 6).number_format = "0.00"
            if leg.expense is not None:
                s2.cell(r, 7).value = float(leg.expense)
                s2.cell(r, 7).number_format = "0.00"
            s2.cell(r, 8).value = f"=SUM(F{r}:G{r})"
            s2.cell(r, 8).number_format = "0.00"
            for col in range(2, 9):
                s2.cell(r, col).border = box
            r += 1

        if len(travels) > 1 and t is not travels[-1]:
            r += 1  # blank row between trips

    last_data_row = r - 1
    r += 1  # blank row before totals

    spolu_row = r
    s2.cell(r, 2).value = "SPOLU"; s2.cell(r, 2).font = bold
    s2.cell(r, 6).value = f"=SUM(F{data_start}:F{last_data_row})"
    s2.cell(r, 6).number_format = "0.00"; s2.cell(r, 6).font = bold
    s2.cell(r, 7).value = f"=SUM(G{data_start}:G{last_data_row})"
    s2.cell(r, 7).number_format = "0.00"; s2.cell(r, 7).font = bold
    s2.cell(r, 8).value = f"=SUM(H{data_start}:H{last_data_row})"
    s2.cell(r, 8).number_format = "0.00"; s2.cell(r, 8).font = bold

    preddavok_row = r + 1
    s2.cell(preddavok_row, 2).value = "PREDDAVOK"
    s2.cell(preddavok_row, 8).value = 0.0
    s2.cell(preddavok_row, 8).number_format = "0.00"

    doplatok_row = r + 2
    s2.cell(doplatok_row, 2).value = "DOPLATOK – PREPLATOK"; s2.cell(doplatok_row, 2).font = bold
    s2.cell(doplatok_row, 8).value = f"=H{spolu_row}-H{preddavok_row}"
    s2.cell(doplatok_row, 8).number_format = "0.00"; s2.cell(doplatok_row, 8).font = bold

    for col, w in {"B": 14, "C": 28, "D": 9, "E": 28, "F": 10, "G": 10, "H": 10}.items():
        s2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
